import keyboard
import requests
import base64
import os
from PIL import ImageGrab
from openpyxl import Workbook, load_workbook
import win32com.client

# Mathpix credentials
APP_ID = "..."
APP_KEY = "..."

# Image save settings
SAVE_DIR = "images"
EXCEL_FILE = "MatematikaPaveiksl.xlsx"
os.makedirs(SAVE_DIR, exist_ok=True)

current_qid = 1702
answer_index = 1
active_row = 6

def get_clipboard_image():
    image = ImageGrab.grabclipboard()
    if image is None:
        print("No image found in clipboard.")
        return None
    return image

def send_to_mathpix(image):
    buffered = image_to_base64(image)
    response = requests.post(
        "https://api.mathpix.com/v3/text",
        headers={
            "app_id": APP_ID,
            "app_key": APP_KEY,
            "Content-type": "application/json"
        },
        json={
            "src": f"data:image/png;base64,{buffered}",
            "formats": ["text", "data"],
            "math_inline_delimiters": ["$$", "$$"],
            "math_display_delimiters": ["$$", "$$"]
        }
    )
    if response.ok:
        result = response.json()
        print("\nMathpix OCR Result:")
        print(result.get("text"))
        return result.get("text")
    else:
        print("Error:", response.text)
        return None

def image_to_base64(image):
    import io
    buffered = io.BytesIO()
    image.save(buffered, format="PNG")
    return base64.b64encode(buffered.getvalue()).decode()

def save_image(filename):
    image = get_clipboard_image()
    if image:
        filepath = os.path.join(SAVE_DIR, filename)
        image.save(filepath)
        print(f"Saved image as: {filepath}")
    else:
        print("Clipboard does not contain an image.")


def log_to_excel(qid, latex, typeLatex):
    global active_row
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "category_id", "question", "correct_answer", "false_answer1", "false_answer2", "false_answer3", "fa_check", "image"])
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active



    match typeLatex:
        case "question":
            ws.cell(row=active_row, column=1).value = qid  # id
            ws.cell(row=active_row, column=3).value = latex  # question
        case "correct_answer":
            ws.cell(row=active_row, column=4).value = latex  # correct_answer
            ws.cell(row=active_row, column=7).value = True  # correct_answer
        case "false_answer":
            ws.cell(row=active_row, column=7).value = False  # correct_answer
            for col in range(5, 8):
                if ws.cell(row=active_row, column=col).value in [None, ""]:
                    ws.cell(row=active_row, column=col).value = latex
                    print(f"🟡 Logged false answer in column {col}")
                    break
        case "image":
            ws.cell(row=active_row, column=9).value = "https://exvpdduqmfmvkvpmbpvp.supabase.co/storage/v1/object/public/task-pictures//" + latex
        case "image_answer":
            ws.cell(row=active_row, column=7).value = True  # correct_answer
            for col in range(5, 8):
                if ws.cell(row=active_row, column=col).value in [None, ""]:
                    ws.cell(row=active_row, column=col).value = "https://exvpdduqmfmvkvpmbpvp.supabase.co/storage/v1/object/public/task-pictures//" + latex
                    print(f"🟡 Logged false answer in column {col}")
                    break

    wb.save(EXCEL_FILE)
    print(f"📗 Logged to Excel: {EXCEL_FILE}")

def ocr_clipboard_image_question_text():
    image = get_clipboard_image()
    if image:
        result = send_to_mathpix(image)
        log_to_excel(current_qid, result, "question")

def ocr_clipboard_image_answer_text():
    image = get_clipboard_image()
    if image:
        result = send_to_mathpix(image)
        log_to_excel(current_qid, result, "correct_answer")

def ocr_clipboard_image_false_answer_text():
    image = get_clipboard_image()
    if image:
        result = send_to_mathpix(image)
        log_to_excel(current_qid, result, "false_answer")

def save_question_image():
    save_image(f"{current_qid}.png")
    log_to_excel(current_qid, f'{current_qid}.png', "image")

def save_answer_image():
    global answer_index
    save_image(f"{current_qid}-{answer_index}.png")
    log_to_excel(current_qid, f'{current_qid}.png', "image_answer")
    answer_index += 1
    if answer_index > 4:
        answer_index = 1

def finalize_question():
    global current_qid, answer_index, active_row
    current_qid += 1
    active_row += 1
    answer_index = 1
    print(f"✅ Finished question. Moving to next: {current_qid}")

# Hotkeys
keyboard.add_hotkey("q", ocr_clipboard_image_question_text)  # OCR question text
keyboard.add_hotkey("a", ocr_clipboard_image_answer_text)  # OCR answer text
keyboard.add_hotkey("f", ocr_clipboard_image_false_answer_text)  # OCR answer text
keyboard.add_hotkey("s", save_question_image)        # Save question image
keyboard.add_hotkey("1", save_answer_image)          # Save answer image
keyboard.add_hotkey("n", finalize_question) #Next question

print("🟢 Tool running. Use Win+Shift+S to snip, then:")
print("🔹 Q → OCR question text via Mathpix")
print("🔹 A → OCR answer text via Mathpix")
print("🔹 F → OCR false answer text via Mathpix")
print("🔹 S → Save image question")
print("🔹 1 → Save as answer image")
print("🔹 N → Next question")
print("🔺 Press ESC to quit.")

keyboard.wait("esc")
